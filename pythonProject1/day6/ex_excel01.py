# pip install openpyxl
from openpyxl import load_workbook
load_wk = load_workbook('./emp.xlsx', data_only=True)
load_wk = load_wk['EMPLOYEES']

# 행 단위로 출력
for row in load_wk.rows:
    print(row)
    for cell in row:
        print(cell.value)
# 셀 주소로 출력
print(load_wk['A1'].value)
# 셀 좌표로 출력
print(load_wk.cell(2, 2).value)
